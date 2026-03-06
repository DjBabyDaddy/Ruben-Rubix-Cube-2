import time
import base64
import threading
import cv2
import glob
import os
from io import BytesIO
from llm import get_llm_multimodal_output
from tts import edge_speak

_TEXT_EXTENSIONS = ['txt', 'md', 'py', 'js', 'ts', 'json', 'csv', 'xml',
                    'yaml', 'yml', 'html', 'css', 'log', 'ini', 'cfg']
_DOC_EXTENSIONS  = _TEXT_EXTENSIONS + ['pdf', 'docx', 'doc', 'xlsx', 'xls']

def get_latest_video():
    """Hunts the PC for the most recent video file."""
    video_dir = os.path.join(os.path.expanduser("~"), "Videos", "Captures")
    if not os.path.exists(video_dir):
        return None
        
    try:
        files = glob.glob(os.path.join(video_dir, '*.mp4')) + glob.glob(os.path.join(video_dir, '*.mkv'))
        if not files: return None
        return max(files, key=os.path.getmtime)
    except Exception as e:
        print(f"⚠️ Media retrieval error: {e}")
        return None

def analyze_multimodal_view(parameters, response, player, session_memory):
    """The central routing hub for RUBE's optical inputs."""
    user_prompt = parameters.get("prompt", "Describe what you see.").strip()
    target = parameters.get("target", "camera").strip().lower()
    file_path = parameters.get("file_path", "").strip()
    
    print(f"👁️ RUBE Vision Matrix Active: Target='{target}', Prompt='{user_prompt}', File='{file_path}'")
    
    if file_path or target == "file":
        threading.Thread(target=_analyze_file, args=(file_path, user_prompt, player), daemon=True).start()
    elif target == "screen":
        threading.Thread(target=_analyze_screen, args=(user_prompt, player), daemon=True).start()
    elif target in ["latest_video", "video", "clip"]:
        threading.Thread(target=_analyze_video, args=(user_prompt, player, None), daemon=True).start()
    else:
        threading.Thread(target=_analyze_camera, args=(user_prompt, player), daemon=True).start()

def _analyze_file(file_path, user_prompt, player):
    print(f"📂 Accessing dropped file: {file_path}")
    file_path = file_path.replace('"', '').strip()
    
    if not os.path.exists(file_path):
        msg = f"Boss, I cannot locate the file at that path."
        edge_speak(msg, player)
        return
        
    ext = file_path.lower().split('.')[-1]
    
    if ext in ['mp4', 'mkv', 'avi', 'mov', 'webm']:
        _analyze_video(user_prompt, player, specific_path=file_path)
    elif ext in ['jpg', 'jpeg', 'png', 'bmp', 'webp']:
        try:
            from PIL import Image
            img = Image.open(file_path)

            img.thumbnail((1024, 1024), Image.Resampling.LANCZOS)

            if img.mode != 'RGB':
                img = img.convert('RGB')

            buffered = BytesIO()
            img.save(buffered, format="JPEG", quality=75)
            img_str = base64.b64encode(buffered.getvalue()).decode('utf-8')

            visual_summary = get_llm_multimodal_output([img_str], user_prompt)
            edge_speak(visual_summary, player)
        except Exception as e:
            print(f"⚠️ IMAGE ERROR: {e}")
            msg = "I encountered a matrix error while trying to parse the image file."
            edge_speak(msg, player)
    elif ext in _DOC_EXTENSIONS:
        _analyze_document(file_path, user_prompt, player)
    else:
        msg = "Boss, that specific file format is currently incompatible with my visual cortex."
        edge_speak(msg, player)


def _analyze_document(file_path, user_prompt, player):
    """Reads text-based documents and routes them through Claude for analysis."""
    ext = file_path.lower().split('.')[-1]
    label = os.path.basename(file_path)
    content = ""

    print(f"📄 RUBE Document Matrix: Reading {label}...")

    try:
        if ext in _TEXT_EXTENSIONS:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()

        elif ext == 'pdf':
            import pdfplumber
            with pdfplumber.open(file_path) as pdf:
                content = "\n".join(page.extract_text() or "" for page in pdf.pages)

        elif ext in ['docx', 'doc']:
            from docx import Document as DocxDocument
            doc = DocxDocument(file_path)
            content = "\n".join(p.text for p in doc.paragraphs if p.text.strip())

        elif ext in ['xlsx', 'xls']:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            rows = []
            for sheet_name in list(wb.sheetnames)[:3]:
                ws = wb[sheet_name]
                rows.append(f"[Sheet: {sheet_name}]")
                for row in ws.iter_rows(max_row=200, values_only=True):
                    line = ", ".join(str(c) for c in row if c is not None)
                    if line.strip():
                        rows.append(line)
            content = "\n".join(rows)

        if not content.strip():
            edge_speak("Boss, I opened the file but it appears to be empty.", player)
            return

        # Truncate at ~15K chars to stay within context limits
        if len(content) > 15000:
            content = content[:15000]
            print(f"⚠️ Document truncated to 15,000 characters.")

        print(f"✅ Document loaded: {len(content)} chars from {label}.")
        full_prompt = f"{user_prompt}\n\n[DOCUMENT — {label}]:\n{content}"
        analysis = get_llm_multimodal_output([], full_prompt)
        edge_speak(analysis, player)

    except Exception as e:
        print(f"⚠️ DOCUMENT ERROR: {e}")
        edge_speak("Boss, I hit an error while trying to read that document.", player)

def _analyze_screen(user_prompt, player):
    print("📸 Scanning user desktop...")
    time.sleep(1.0) 
    try:
        from PIL import ImageGrab
        screenshot = ImageGrab.grab()
        buffered = BytesIO()
        screenshot.save(buffered, format="JPEG", quality=80)
        img_str = base64.b64encode(buffered.getvalue()).decode("utf-8")
        
        visual_summary = get_llm_multimodal_output([img_str], user_prompt)
        edge_speak(visual_summary, player)
    except Exception as e:
        print(f"⚠️ SCREENSHOT ERROR: {e}")
        msg = "I encountered an error trying to capture your screen, boss."
        edge_speak(msg, player)

def _analyze_camera(user_prompt, player):
    print("📷 Accessing webcam feed...")
    time.sleep(1.2) 
    try:
        cam = cv2.VideoCapture(0)
        if not cam.isOpened():
            msg = "Boss, I cannot activate the visual matrix. The webcam feed appears disconnected."
            edge_speak(msg, player)
            return
            
        ret, frame = cam.read()
        cam.release() 
        
        if not ret:
            msg = "Boss, I captured a corrupt visual frame."
            edge_speak(msg, player)
            return
            
        _, buffer = cv2.imencode('.jpg', frame, [int(cv2.IMWRITE_JPEG_QUALITY), 80])
        image_base64 = base64.b64encode(buffer).decode('utf-8')
        
        visual_summary = get_llm_multimodal_output([image_base64], user_prompt)
        edge_speak(visual_summary, player)
    except Exception as e:
        print(f"⚠️ CAMERA ERROR: {e}")
        msg = "My camera module crashed during execution, boss."
        edge_speak(msg, player)

def _get_frame_count(total_frames, fps):
    """Scale frame extraction based on video duration for accurate analysis."""
    duration = (total_frames / fps) if fps > 0 else 0
    if duration < 30:    return 6
    if duration < 300:   return 12   # up to 5 min
    if duration < 1800:  return 20   # up to 30 min
    return 30                        # 30+ min


def _analyze_video(user_prompt, player, specific_path=None):
    print("🎞️ Ripping video frames for storyboard analysis...")
    video_path = specific_path if specific_path else get_latest_video()

    if not video_path:
        msg = "Boss, I could not find any recent videos to analyze."
        edge_speak(msg, player)
        return

    try:
        cam = cv2.VideoCapture(video_path)
        total_frames = int(cam.get(cv2.CAP_PROP_FRAME_COUNT))
        fps = cam.get(cv2.CAP_PROP_FPS)

        frames_to_grab = _get_frame_count(total_frames, fps)
        step = max(1, total_frames // frames_to_grab)

        base64_frames = []
        for i in range(frames_to_grab):
            cam.set(cv2.CAP_PROP_POS_FRAMES, i * step)
            ret, frame = cam.read()
            if ret:
                frame = cv2.resize(frame, (1280, 720))
                _, buffer = cv2.imencode('.jpg', frame, [int(cv2.IMWRITE_JPEG_QUALITY), 75])
                base64_frames.append(base64.b64encode(buffer).decode('utf-8'))

        cam.release()
        print(f"✅ Extracted {len(base64_frames)} frames from {os.path.basename(video_path)}.")

        prompt_override = f"I have extracted {len(base64_frames)} sequential frames from a video. Based on the progression of these images, {user_prompt}"
        visual_summary = get_llm_multimodal_output(base64_frames, prompt_override)
        
        edge_speak(visual_summary, player)
    except Exception as e:
        print(f"⚠️ VIDEO PROCESSING ERROR: {e}")
        msg = "My visual matrix crashed while trying to parse the video file."
        edge_speak(msg, player)